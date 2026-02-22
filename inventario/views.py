from django.shortcuts import render, redirect
from django.template.loader import get_template
from django.http import HttpResponse, JsonResponse
from django.urls import reverse_lazy
from django.contrib.auth.views import LoginView
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.hashers import make_password
from django.db.models import Sum, Count, Q, F
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from datetime import timedelta, datetime
from django.db import transaction
from collections import defaultdict
import json
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from materiales.models import material
from inventarios.models import Inventarios, AlertaStock, stock_inv
from usuarios.models import CustomUser
from transacciones.models import Movimientos

# ============================================
# VISTA DE LOGIN
# ============================================
class Login_View(LoginView):
    template_name = "login.html"
    redirect_authenticated_user = True

    def form_invalid(self, form):
        print("Errores del formulario:", form.errors)
        messages.error(self.request, 'Usuario o contrasena incorrectos')
        return super().form_invalid(form)

    def get_success_url(self):
        user = self.request.user
        if hasattr(user, 'rol'):
            if user.rol == CustomUser.Rol.ADMIN:
                return reverse_lazy('dashboard')
            elif user.rol == CustomUser.Rol.ESCUELA:
                return reverse_lazy('dashboard')
        return self.request.GET.get('next', 'dashboard')


# ============================================
# VISTA HOME
# ============================================
def home(request):
    user = request.user

    if user.is_authenticated:
        return redirect('dashboard')

    return redirect('login')


# ============================================
# VISTA DASHBOARD
# ============================================
@login_required
def dashboard(request):
    user = request.user
    inv = user.admin_of if hasattr(user, 'admin_of') else None

    # ===== OBTENER DATOS CON OPTIMIZACION =====
    materiales = material.objects.all().prefetch_related('stock_inv_set__inv')
    escuelas = Inventarios.objects.filter(es_depo=False)
    depositos = Inventarios.objects.filter(es_depo=True)
    alertas = AlertaStock.objects.filter(resuelto=False).select_related(
        'inv', 'mat'
    ).order_by('-fecha')
    alertas_recientes = alertas[:10]
    stocks = stock_inv.objects.select_related('inv', 'mat').all()
    usuarios = CustomUser.objects.all().select_related('admin_of')

    # ===== CALCULAR ESTADISTICAS =====
    total_escuelas = escuelas.count()
    total_depositos = depositos.count()
    total_materiales = materiales.count()
    total_usuarios = usuarios.count()
    total_items_stock = stocks.aggregate(total=Sum('mat_qty'))['total'] or 0
    alertas_criticas = alertas.count()
    pedidos_pendientes = 0

    # ===== CLASIFICAR MATERIALES POR TIPO =====
    materiales_por_tipo = defaultdict(int)
    for mat in materiales:
        materiales_por_tipo[mat.tipo] += 1
    materiales_por_tipo = dict(materiales_por_tipo)

    # ===== ESTADISTICAS DE USUARIOS =====
    usuarios_administradores = usuarios.filter(
        rol=CustomUser.Rol.ADMIN
    ).count()
    usuarios_escuelas = usuarios.filter(rol=CustomUser.Rol.ESCUELA).count()
    usuarios_inactivos = usuarios.filter(is_active=False).count()
    usuarios_activos = usuarios.filter(is_active=True).count()

    # ===== PREPARAR DICCIONARIO DE STOCK POR MATERIAL =====
    stock_por_material = {}
    stock_minimo_por_material = {}

    for s in stocks:
        if s.mat_id not in stock_por_material:
            stock_por_material[s.mat_id] = 0
            stock_minimo_por_material[s.mat_id] = []
        stock_por_material[s.mat_id] += s.mat_qty
        stock_minimo_por_material[s.mat_id].append(s.mat_min_qty)

    for mat_id in stock_minimo_por_material:
        if stock_minimo_por_material[mat_id]:
            stock_minimo_por_material[mat_id] = (
                sum(stock_minimo_por_material[mat_id])
                / len(stock_minimo_por_material[mat_id])
            )
        else:
            stock_minimo_por_material[mat_id] = 0

    # ===== PREPARAR DATOS PARA JAVASCRIPT =====
    escuelas_json = []
    for escuela in escuelas:
        usuarios_count = CustomUser.objects.filter(admin_of=escuela).count()
        stocks_escuela = stock_inv.objects.filter(inv=escuela)
        materiales_count = (
            stocks_escuela.aggregate(total=Sum('mat_qty'))['total'] or 0
        )

        escuelas_json.append(
            {
                'id': escuela.id,
                'name': escuela.name,
                'ubic': escuela.ubic or '',
                'usuarios_count': usuarios_count,
                'materiales_count': materiales_count,
                'es_depo': escuela.es_depo,
                'activo': True,
            }
        )

    materiales_json = []
    for mat in materiales:
        stock_total = stock_por_material.get(mat.id, 0)
        stock_minimo = stock_minimo_por_material.get(mat.id, 5)

        if stock_total == 0:
            status_class = 'out-of-stock'
            status_text = 'Agotado'
        elif stock_total < stock_minimo:
            status_class = 'low-stock'
            status_text = 'Bajo Stock'
        else:
            status_class = 'in-stock'
            status_text = 'En Stock'

        materiales_json.append(
            {
                'id': mat.id,
                'name': mat.name,
                'desc': mat.desc or '',
                'tipo': mat.tipo,
                'tipo_display': mat.get_tipo_display(),
                'unidad': mat.get_unidad_display(),
                'stock_total': stock_total,
                'stock_minimo': stock_minimo,
                'status_class': status_class,
                'status_text': status_text,
            }
        )

    usuarios_json = []
    for usr in usuarios:
        escuela_nombre = usr.admin_of.name if usr.admin_of else 'Sistema'

        usuarios_json.append(
            {
                'id': usr.id,
                'username': usr.username,
                'first_name': usr.first_name or '',
                'last_name': usr.last_name or '',
                'full_name': f"{usr.first_name} {usr.last_name}".strip(),
                'email': usr.email or '',
                'rol': usr.rol,
                'rol_display': usr.get_rol_display(),
                'nro_tlf': usr.nro_tlf,
                'escuela_id': usr.admin_of_id,
                'escuela_nombre': escuela_nombre,
                'is_active': usr.is_active,
                'last_login': usr.last_login.isoformat()
                if usr.last_login
                else None,
            }
        )

    alertas_json = []
    for alerta in alertas_recientes:
        alertas_json.append(
            {
                'id': alerta.id,
                'msg': alerta.msg,
                'fecha': alerta.fecha.isoformat(),
                'inventario_id': alerta.inv_id,
                'escuela_nombre': alerta.inv.name if alerta.inv else 'Desconocida',
                'material_nombre': alerta.mat.name if alerta.mat else 'Material',
                'resuelto': alerta.resuelto,
            }
        )

    stocks_por_escuela = []
    for escuela in escuelas:
        stocks_escuela = stock_inv.objects.filter(inv=escuela).select_related(
            'mat'
        )
        items = []
        for s in stocks_escuela:
            items.append(
                {
                    'material_id': s.mat_id,
                    'material_nombre': s.mat.name,
                    'cantidad': s.mat_qty,
                    'minimo': s.mat_min_qty,
                    'estado': 'Agotado'
                    if s.mat_qty == 0
                    else (
                        'Bajo Stock'
                        if s.mat_qty < s.mat_min_qty
                        else 'Normal'
                    ),
                }
            )

        stocks_por_escuela.append(
            {
                'escuela_id': escuela.id,
                'escuela_nombre': escuela.name,
                'items': items,
            }
        )

    data_json = {
        'escuelas': escuelas_json,
        'materiales': materiales_json,
        'usuarios': usuarios_json,
        'alertas': alertas_json,
        'stocks_por_escuela': stocks_por_escuela,
        'estadisticas': {
            'total_escuelas': total_escuelas,
            'total_depositos': total_depositos,
            'total_materiales': total_materiales,
            'total_usuarios': total_usuarios,
            'total_items_stock': total_items_stock,
            'alertas_criticas': alertas_criticas,
            'pedidos_pendientes': pedidos_pendientes,
            'usuarios_activos': usuarios_activos,
            'usuarios_inactivos': usuarios_inactivos,
            'usuarios_administradores': usuarios_administradores,
            'usuarios_escuelas': usuarios_escuelas,
        },
    }

    context = {
        "user": user,
        "materiales": materiales,
        "materiales_json": json.dumps(materiales_json),
        "alertas": alertas,
        "recent_alerts": alertas_recientes,
        "inventarios": inv,
        "escuelas": escuelas,
        "depositos": depositos,
        "stocks": stocks,
        "usuarios": usuarios,
        "total_escuelas": total_escuelas,
        "total_materiales": total_materiales,
        "total_items_stock": total_items_stock,
        "alertas_criticas": alertas_criticas,
        "materiales_por_tipo": materiales_por_tipo,
        "usuarios_administradores": usuarios_administradores,
        "usuarios_escuelas": usuarios_escuelas,
        "usuarios_activos": usuarios_activos,
        "usuarios_inactivos": usuarios_inactivos,
        "CustomUser": CustomUser,
        "material_model": material,
        "data_json": json.dumps(data_json),
    }

    if hasattr(user, 'rol'):
        if user.rol == CustomUser.Rol.ADMIN:
            return render(request, "dashboard.html", context)
        elif user.rol == CustomUser.Rol.ESCUELA:
            escuela_usuario = user.admin_of
            if escuela_usuario:
                stocks_escuela = stock_inv.objects.filter(
                    inv=escuela_usuario
                ).select_related('mat')
                materiales_escuela = [s.mat for s in stocks_escuela]
                alertas_escuela = alertas.filter(inv=escuela_usuario)
                total_stock = (
                    stocks_escuela.aggregate(total=Sum('mat_qty'))['total']
                    or 0
                )
            else:
                stocks_escuela = []
                materiales_escuela = []
                alertas_escuela = alertas.none()
                total_stock = 0

            return render(
                request,
                "dashboardes.html",
                {
                    "user": user,
                    "materiales": materiales_escuela,
                    "stocks": stocks_escuela,
                    "alertas": alertas_escuela,
                    "recent_alerts": alertas_escuela[:5],
                    "total_stock": total_stock,
                    "escuela": escuela_usuario,
                },
            )

    return redirect('login')


# ============================================
# VISTA NOSOTROS
# ============================================
def nosotros(request):
    plantilla = get_template("nosotros.html")
    return HttpResponse(plantilla.render())


# ============================================
# VISTA DE LOGOUT
# ============================================
def logout_view(request):
    logout(request)
    messages.success(request, 'Sesion cerrada correctamente')
    return redirect('login')


# ============================================
# API ENDPOINTS - MATERIALES
# ============================================

@login_required
def api_tipos_material(request):
    """Endpoint para obtener los tipos de material del modelo"""
    tipos = []
    for tipo_value, tipo_label in material.Tipo.choices:
        tipos.append({'valor': tipo_value, 'etiqueta': tipo_label})
    return JsonResponse({'tipos': tipos})


@login_required
def api_unidades_material(request):
    """Endpoint para obtener las unidades de medida del modelo"""
    unidades = []
    for unidad_value, unidad_label in material.Unidad.choices:
        unidades.append({'valor': unidad_value, 'etiqueta': unidad_label})
    return JsonResponse({'unidades': unidades})


@login_required
def api_materiales(request):
    """Listar y crear materiales"""
    if request.method == 'GET':
        materiales = material.objects.all()
        data = []
        for mat in materiales:
            stock_total = (
                stock_inv.objects.filter(mat=mat).aggregate(total=Sum('mat_qty'))[
                    'total'
                ]
                or 0
            )
            stock_minimo = (
                stock_inv.objects.filter(mat=mat).aggregate(
                    avg=Sum('mat_min_qty') / Count('id')
                )['avg']
                or 5
            )

            data.append(
                {
                    'id': mat.id,
                    'name': mat.name,
                    'desc': mat.desc,
                    'tipo': mat.tipo,
                    'tipo_display': mat.get_tipo_display(),
                    'unidad': mat.unidad,
                    'unidad_display': mat.get_unidad_display(),
                    'stock_total': stock_total,
                    'stock_minimo': stock_minimo,
                }
            )
        return JsonResponse({'data': data})

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)

            if material.objects.filter(name=data['name']).exists():
                return JsonResponse(
                    {'error': 'Ya existe un material con ese nombre'}, status=400
                )

            mat = material.objects.create(
                name=data['name'],
                tipo=data['tipo'],
                unidad=data.get('unidad', 'PIEZAS'),
                desc=data.get('desc', ''),
            )

            return JsonResponse(
                {'message': 'Material creado correctamente', 'id': mat.id}
            )
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_material_detalle(request, material_id):
    """Obtener, actualizar o eliminar un material especifico"""
    try:
        mat = material.objects.get(id=material_id)
    except material.DoesNotExist:
        return JsonResponse({'error': 'Material no encontrado'}, status=404)

    if request.method == 'GET':
        stock_total = (
            stock_inv.objects.filter(mat=mat).aggregate(total=Sum('mat_qty'))['total']
            or 0
        )
        return JsonResponse(
            {
                'id': mat.id,
                'name': mat.name,
                'desc': mat.desc,
                'tipo': mat.tipo,
                'unidad': mat.unidad,
                'stock_total': stock_total,
            }
        )

    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)

            if (
                material.objects.filter(name=data['name'])
                .exclude(id=material_id)
                .exists()
            ):
                return JsonResponse(
                    {'error': 'Ya existe otro material con ese nombre'}, status=400
                )

            mat.name = data.get('name', mat.name)
            mat.tipo = data.get('tipo', mat.tipo)
            mat.unidad = data.get('unidad', mat.unidad)
            mat.desc = data.get('desc', mat.desc)
            mat.save()

            return JsonResponse({'message': 'Material actualizado correctamente'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    elif request.method == 'DELETE':
        try:
            if stock_inv.objects.filter(mat=mat).exists():
                return JsonResponse(
                    {'error': 'No se puede eliminar porque tiene stocks asociados'},
                    status=400,
                )

            mat.delete()
            return JsonResponse({'message': 'Material eliminado correctamente'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_filtrar_materiales(request):
    """Filtrar materiales por busqueda, tipo y escuela"""
    search = request.GET.get('search', '')
    tipo = request.GET.get('tipo', '')
    escuela_id = request.GET.get('escuela', '')

    materiales_qs = material.objects.all()

    if search:
        materiales_qs = materiales_qs.filter(
            Q(name__icontains=search) | Q(desc__icontains=search)
        )

    if tipo:
        materiales_qs = materiales_qs.filter(tipo=tipo)

    stocks_qs = stock_inv.objects.all()
    if escuela_id:
        stocks_qs = stocks_qs.filter(inv_id=escuela_id)

    stock_dict = {}
    for s in stocks_qs:
        if s.mat_id not in stock_dict:
            stock_dict[s.mat_id] = 0
        stock_dict[s.mat_id] += s.mat_qty

    resultados = []
    for mat in materiales_qs:
        stock_total = stock_dict.get(mat.id, 0)

        if escuela_id and stock_total == 0:
            continue

        resultados.append(
            {
                'id': mat.id,
                'name': mat.name,
                'desc': mat.desc,
                'tipo': mat.tipo,
                'tipo_display': mat.get_tipo_display(),
                'unidad': mat.unidad,
                'stock_total': stock_total,
            }
        )

    return JsonResponse({'data': resultados})


@login_required
def api_exportar_inventario(request):
    """Exportar inventario a Excel o CSV"""
    export_format = request.GET.get('format', 'csv')

    materiales = material.objects.all()
    stocks = stock_inv.objects.select_related('inv', 'mat').all()

    stock_data = {}
    for s in stocks:
        key = (s.mat_id, s.inv_id)
        stock_data[key] = s

    if export_format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        headers = [
            'ID',
            'Material',
            'Tipo',
            'Unidad',
            'Descripcion',
            'Escuela',
            'Cantidad',
            'Minimo',
            'Estado',
        ]
        ws.append(headers)

        for mat in materiales:
            for escuela in Inventarios.objects.filter(es_depo=False):
                key = (mat.id, escuela.id)
                s = stock_data.get(key)

                cantidad = s.mat_qty if s else 0
                minimo = s.mat_min_qty if s else 5

                if cantidad == 0:
                    estado = 'Agotado'
                elif cantidad < minimo:
                    estado = 'Bajo Stock'
                else:
                    estado = 'Normal'

                ws.append(
                    [
                        mat.id,
                        mat.name,
                        mat.get_tipo_display(),
                        mat.get_unidad_display(),
                        mat.desc or '',
                        escuela.name,
                        cantidad,
                        minimo,
                        estado,
                    ]
                )

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response[
            'Content-Disposition'
        ] = f'attachment; filename=inventario_{request.GET.get("date", "")}.xlsx'
        wb.save(response)
        return response

    else:
        response = HttpResponse(content_type='text/csv')
        response[
            'Content-Disposition'
        ] = f'attachment; filename=inventario_{request.GET.get("date", "")}.csv'

        writer = csv.writer(response)
        writer.writerow(
            [
                'ID',
                'Material',
                'Tipo',
                'Unidad',
                'Descripcion',
                'Escuela',
                'Cantidad',
                'Minimo',
                'Estado',
            ]
        )

        for mat in materiales:
            for escuela in Inventarios.objects.filter(es_depo=False):
                key = (mat.id, escuela.id)
                s = stock_data.get(key)

                cantidad = s.mat_qty if s else 0
                minimo = s.mat_min_qty if s else 5

                if cantidad == 0:
                    estado = 'Agotado'
                elif cantidad < minimo:
                    estado = 'Bajo Stock'
                else:
                    estado = 'Normal'

                writer.writerow(
                    [
                        mat.id,
                        mat.name,
                        mat.get_tipo_display(),
                        mat.get_unidad_display(),
                        mat.desc or '',
                        escuela.name,
                        cantidad,
                        minimo,
                        estado,
                    ]
                )

        return response


@login_required
def api_stocks(request):
    """Endpoint API para obtener todos los stocks"""
    try:
        stocks = stock_inv.objects.select_related('inv', 'mat').all()

        data = []
        for s in stocks:
            if s.mat_qty == 0:
                estado = 'Agotado'
            elif s.mat_qty < s.mat_min_qty:
                estado = 'Bajo Stock'
            else:
                estado = 'Normal'

            data.append(
                {
                    'id': s.id,
                    'inventario_id': s.inv_id,
                    'inventario_nombre': s.inv.name if s.inv else 'Desconocido',
                    'material_id': s.mat_id,
                    'material_nombre': s.mat.name if s.mat else 'Desconocido',
                    'cantidad': s.mat_qty,
                    'minimo': s.mat_min_qty,
                    'estado': estado,
                    'ubicacion': s.inv.ubic if s.inv else 'No especificada',
                }
            )

        return JsonResponse({'data': data, 'total': len(data)})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_stocks_por_escuela(request, escuela_id):
    """Endpoint API para obtener stocks de una escuela especifica"""
    try:
        stocks = stock_inv.objects.filter(inv_id=escuela_id).select_related(
            'inv', 'mat'
        )

        data = []
        for s in stocks:
            data.append(
                {
                    'id': s.id,
                    'material_id': s.mat_id,
                    'material_nombre': s.mat.name,
                    'cantidad': s.mat_qty,
                    'minimo': s.mat_min_qty,
                    'estado': 'Agotado'
                    if s.mat_qty == 0
                    else (
                        'Bajo Stock' if s.mat_qty < s.mat_min_qty else 'Normal'
                    ),
                }
            )

        return JsonResponse({'data': data})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_stock_detalle(request, stock_id):
    """Obtener, actualizar o eliminar un stock especifico"""
    try:
        stock = stock_inv.objects.select_related('inv', 'mat').get(id=stock_id)
    except stock_inv.DoesNotExist:
        return JsonResponse({'error': 'Stock no encontrado'}, status=404)

    if request.method == 'GET':
        return JsonResponse(
            {
                'id': stock.id,
                'inventario_id': stock.inv_id,
                'material_id': stock.mat_id,
                'cantidad': stock.mat_qty,
                'minimo': stock.mat_min_qty,
            }
        )

    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)

            stock.mat_qty = data.get('cantidad', stock.mat_qty)
            stock.mat_min_qty = data.get('minimo', stock.mat_min_qty)
            stock.save()

            return JsonResponse({'message': 'Stock actualizado correctamente'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    elif request.method == 'DELETE':
        try:
            stock.delete()
            return JsonResponse({'message': 'Stock eliminado correctamente'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


# ============================================
# API ENDPOINTS - USUARIOS
# ============================================

@login_required
def api_usuarios(request):
    """Listar y crear usuarios"""
    if request.method == 'GET':
        usuarios = CustomUser.objects.all().select_related('admin_of')
        data = []
        for usr in usuarios:
            data.append(
                {
                    'id': usr.id,
                    'username': usr.username,
                    'first_name': usr.first_name,
                    'last_name': usr.last_name,
                    'email': usr.email,
                    'rol': usr.rol,
                    'rol_display': usr.get_rol_display(),
                    'escuela_id': usr.admin_of_id,
                    'escuela_nombre': usr.admin_of.name if usr.admin_of else None,
                    'nro_tlf': usr.nro_tlf,
                    'is_active': usr.is_active,
                    'last_login': usr.last_login,
                }
            )
        return JsonResponse({'data': data})

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)

            user = CustomUser.objects.create_user(
                username=data['username'],
                email=data['email'],
                password=data.get('password', ''),
                first_name=data.get('first_name', ''),
                last_name=data.get('last_name', ''),
                rol=data['rol'],
                nro_tlf=data.get('nro_tlf'),
                is_active=data.get('is_active', True),
            )

            if data.get('admin_of'):
                user.admin_of_id = data['admin_of']
                user.save()

            return JsonResponse(
                {'message': 'Usuario creado correctamente', 'id': user.id}
            )
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_usuario_detalle(request, user_id):
    """Obtener, actualizar o eliminar un usuario especifico"""
    try:
        user = CustomUser.objects.get(id=user_id)
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)

    if request.method == 'GET':
        return JsonResponse(
            {
                'id': user.id,
                'username': user.username,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email,
                'rol': user.rol,
                'admin_of': user.admin_of_id,
                'nro_tlf': user.nro_tlf,
                'is_active': user.is_active,
            }
        )

    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)

            user.username = data.get('username', user.username)
            user.first_name = data.get('first_name', user.first_name)
            user.last_name = data.get('last_name', user.last_name)
            user.email = data.get('email', user.email)
            user.rol = data.get('rol', user.rol)
            user.admin_of_id = data.get('admin_of')
            user.nro_tlf = data.get('nro_tlf', user.nro_tlf)
            user.is_active = data.get('is_active', user.is_active)

            if data.get('password'):
                user.password = make_password(data['password'])

            user.save()

            return JsonResponse({'message': 'Usuario actualizado correctamente'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    elif request.method == 'DELETE':
        try:
            user.delete()
            return JsonResponse({'message': 'Usuario eliminado correctamente'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_usuario_toggle_status(request, user_id):
    """Activar o desactivar usuario"""
    try:
        user = CustomUser.objects.get(id=user_id)
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            activate = data.get('activate', True)

            user.is_active = activate
            user.save()

            action = 'activado' if activate else 'desactivado'
            return JsonResponse({'message': f'Usuario {action} correctamente'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_filtrar_usuarios(request):
    """Endpoint para filtrar usuarios en tiempo real"""
    rol = request.GET.get('rol', '')
    escuela = request.GET.get('escuela', '')
    estado = request.GET.get('estado', '')
    busqueda = request.GET.get('busqueda', '').lower()

    usuarios_qs = CustomUser.objects.all().select_related('admin_of')

    if rol:
        usuarios_qs = usuarios_qs.filter(rol=rol)

    if escuela:
        usuarios_qs = usuarios_qs.filter(admin_of_id=escuela)

    if estado:
        is_active = estado == 'true'
        usuarios_qs = usuarios_qs.filter(is_active=is_active)

    if busqueda:
        usuarios_qs = usuarios_qs.filter(
            Q(username__icontains=busqueda)
            | Q(first_name__icontains=busqueda)
            | Q(last_name__icontains=busqueda)
            | Q(email__icontains=busqueda)
        )

    resultados = []
    for usr in usuarios_qs:
        resultados.append(
            {
                'id': usr.id,
                'username': usr.username,
                'first_name': usr.first_name,
                'last_name': usr.last_name,
                'email': usr.email,
                'rol': usr.get_rol_display(),
                'escuela': usr.admin_of.name if usr.admin_of else 'Sistema',
                'telefono': usr.nro_tlf,
                'activo': usr.is_active,
                'ultimo_acceso': usr.last_login.strftime('%d/%m/%Y %H:%M')
                if usr.last_login
                else 'Nunca',
            }
        )

    return JsonResponse({'data': resultados})


# ============================================
# API ENDPOINTS - ESCUELAS
# ============================================

@login_required
def api_filtrar_escuelas(request):
    """Endpoint para filtrar escuelas"""
    busqueda = request.GET.get('busqueda', '').lower()

    escuelas_qs = Inventarios.objects.filter(es_depo=False)

    if busqueda:
        escuelas_qs = escuelas_qs.filter(
            Q(name__icontains=busqueda) | Q(ubic__icontains=busqueda)
        )

    resultados = []
    for escuela in escuelas_qs:
        usuarios_count = CustomUser.objects.filter(admin_of=escuela).count()
        stocks = stock_inv.objects.filter(inv=escuela)
        materiales_count = stocks.count()

        resultados.append(
            {
                'id': escuela.id,
                'name': escuela.name,
                'ubic': escuela.ubic or 'No especificada',
                'usuarios': usuarios_count,
                'materiales': materiales_count,
                'es_depo': escuela.es_depo,
            }
        )

    return JsonResponse({'data': resultados})


# ============================================
# API ENDPOINTS - ALERTAS ESCUELA
# ============================================

@login_required
def api_alertas_escuela(request):
    """Endpoint para obtener alertas de la escuela del usuario"""
    try:
        user = request.user

        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ESCUELA:
            return JsonResponse({'error': 'No autorizado'}, status=403)

        escuela = user.admin_of
        if not escuela:
            return JsonResponse({'data': []})

        alertas = AlertaStock.objects.filter(
            inv=escuela, resuelto=False
        ).select_related('mat').order_by('-fecha')[:10]

        data = []
        for a in alertas:
            data.append(
                {
                    'id': a.id,
                    'material': a.mat.name if a.mat else 'Desconocido',
                    'mensaje': a.msg,
                    'fecha': a.fecha.isoformat() if a.fecha else None,
                }
            )

        return JsonResponse({'data': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ============================================
# API ENDPOINTS - PEDIDOS ADMIN (UNICA VERSION CON DATOS REALES)
# ============================================

@login_required
def api_pedidos_admin(request):
    """Obtener TODOS los movimientos de tipo Salida (pedidos reales)"""
    try:
        user = request.user

        # Verificar que sea admin
        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ADMIN:
            return JsonResponse({'error': 'No autorizado'}, status=403)

        # Obtener TODOS los movimientos de tipo Salida de la base de datos
        pedidos = Movimientos.objects.filter(
            tipo=Movimientos.Tipo.OUT  # 'Salida'
        ).select_related('material', 'user', 'inv_orig', 'inv_dest').order_by(
            '-fecha_Mov'
        )

        print(f"📦 Pedidos encontrados en DB: {pedidos.count()}")  # Para depuracion

        data = []
        for p in pedidos:
            data.append(
                {
                    'id': p.id,
                    'fecha': p.fecha_Mov.isoformat() if p.fecha_Mov else None,
                    'material_id': p.material.id if p.material else None,
                    'material_nombre': p.material.name
                    if p.material
                    else 'Material eliminado',
                    'cantidad': float(p.cant_mov) if p.cant_mov else 0,
                    'unidad': p.material.get_unidad_display() if p.material else '',
                    'escuela_origen': p.inv_orig.name
                    if p.inv_orig
                    else 'Deposito Central',
                    'escuela_destino': p.inv_dest.name if p.inv_dest else None,
                    'solicitante': p.user.username if p.user else 'Sistema',
                    'solicitante_nombre': f"{p.user.first_name} {p.user.last_name}".strip()
                    if p.user and (p.user.first_name or p.user.last_name)
                    else (p.user.username if p.user else 'Sistema'),
                    'estado': getattr(p, 'estado', 'pendiente'),
                    'estado_display': getattr(p, 'get_estado_display', lambda: 'Pendiente')()
                    if hasattr(p, 'get_estado_display')
                    else 'Pendiente',
                }
            )

        # Estadisticas basadas en datos reales
        stats = {
            'pendientes': len([d for d in data if d['estado'] == 'pendiente']),
            'aprobados': len([d for d in data if d['estado'] == 'aprobado']),
            'rechazados': len([d for d in data if d['estado'] == 'rechazado']),
            'entregados': len([d for d in data if d['estado'] == 'entregado']),
            'total': len(data),
        }

        return JsonResponse({'success': True, 'data': data, 'stats': stats})

    except Exception as e:
        print(f"❌ Error en api_pedidos_admin: {str(e)}")
        import traceback

        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
def api_aprobar_pedido(request, pedido_id):
    """Aprobar o rechazar un pedido"""
    try:
        if request.method != 'POST':
            return JsonResponse({'error': 'Metodo no permitido'}, status=405)

        user = request.user

        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ADMIN:
            return JsonResponse({'error': 'No autorizado'}, status=403)

        data = json.loads(request.body)
        accion = data.get('accion')

        try:
            pedido = Movimientos.objects.get(id=pedido_id, tipo='Salida')
        except Movimientos.DoesNotExist:
            return JsonResponse({'error': 'Pedido no encontrado'}, status=404)

        if hasattr(pedido, 'estado'):
            if accion == 'aprobar':
                pedido.estado = 'aprobado'
            elif accion == 'rechazar':
                pedido.estado = 'rechazado'
            elif accion == 'entregar':
                pedido.estado = 'entregado'
            else:
                return JsonResponse({'error': 'Accion no valida'}, status=400)

            pedido.save()

            return JsonResponse(
                {'success': True, 'message': f'Pedido {accion} correctamente'}
            )
        else:
            return JsonResponse(
                {
                    'success': True,
                    'message': f'Pedido {accion} correctamente (simulado)',
                }
            )

    except Exception as e:
        print(f"Error en api_aprobar_pedido: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_estadisticas_pedidos(request):
    """Estadisticas de pedidos"""
    try:
        user = request.user

        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ADMIN:
            return JsonResponse({'error': 'No autorizado'}, status=403)

        total_pedidos = Movimientos.objects.filter(tipo='Salida').count()

        pendientes = Movimientos.objects.filter(
            tipo='Salida', estado='pendiente'
        ).count()
        aprobados = Movimientos.objects.filter(
            tipo='Salida', estado='aprobado'
        ).count()
        rechazados = Movimientos.objects.filter(
            tipo='Salida', estado='rechazado'
        ).count()
        entregados = Movimientos.objects.filter(
            tipo='Salida', estado='entregado'
        ).count()

        top_materiales = (
            Movimientos.objects.filter(tipo='Salida')
            .values('material__name')
            .annotate(total=Count('id'), cantidad_total=Sum('cant_mov'))
            .order_by('-total')[:10]
        )

        return JsonResponse(
            {
                'success': True,
                'estadisticas': {
                    'pendientes': pendientes,
                    'aprobados': aprobados,
                    'rechazados': rechazados,
                    'entregados': entregados,
                    'total': total_pedidos,
                },
                'top_materiales': list(top_materiales),
            }
        )

    except Exception as e:
        print(f"Error en api_estadisticas_pedidos: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# ============================================
# API ENDPOINTS - MOVIMIENTOS ESCUELA
# ============================================

@login_required
def api_movimientos_escuela(request):
    """Endpoint para obtener movimientos/pedidos de la escuela del usuario"""
    try:
        user = request.user

        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ESCUELA:
            return JsonResponse({'error': 'No autorizado'}, status=403)

        escuela = user.admin_of
        if not escuela:
            return JsonResponse({'data': []})

        movimientos = Movimientos.objects.filter(
            Q(inv_orig=escuela) | Q(inv_dest=escuela)
        ).select_related('material', 'user').order_by('-fecha_Mov')

        data = []
        for m in movimientos:
            data.append(
                {
                    'id': m.id,
                    'material_id': m.material.id,
                    'material_nombre': m.material.name,
                    'cantidad': m.cant_mov,
                    'tipo': m.tipo,
                    'fecha': m.fecha_Mov.isoformat(),
                    'origen': m.inv_orig.name if m.inv_orig else None,
                    'destino': m.inv_dest.name if m.inv_dest else None,
                    'usuario': m.user.username if m.user else 'Sistema',
                    'es_salida': m.tipo == Movimientos.Tipo.OUT,
                    'es_entrada': m.tipo == Movimientos.Tipo.IN,
                    'es_ajuste': m.tipo == Movimientos.Tipo.ADJUST,
                }
            )

        return JsonResponse({'data': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_crear_solicitud(request):
    """Endpoint para crear una solicitud de material (OUT + IN)"""
    try:
        if request.method != 'POST':
            return JsonResponse({'error': 'Metodo no permitido'}, status=405)

        user = request.user

        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ESCUELA:
            return JsonResponse({'error': 'No autorizado'}, status=403)

        escuela = user.admin_of
        if not escuela:
            return JsonResponse(
                {'error': 'No tienes una escuela asignada'}, status=400
            )

        data = json.loads(request.body)

        material_id = data.get('material_id')
        cantidad = data.get('cantidad')

        if not material_id or not cantidad:
            return JsonResponse(
                {'error': 'Material y cantidad son requeridos'}, status=400
            )

        try:
            material_obj = material.objects.get(id=material_id)
        except material.DoesNotExist:
            return JsonResponse({'error': 'Material no encontrado'}, status=404)

        deposito_central = Inventarios.objects.filter(es_depo=True).first()
        if not deposito_central:
            return JsonResponse(
                {'error': 'No hay deposito central configurado'}, status=400
            )

        with transaction.atomic():
            stocks_deposito = stock_inv.objects.filter(
                inv=deposito_central, mat=material_obj
            )

            if not stocks_deposito.exists():
                return JsonResponse(
                    {
                        'error': f'El material {material_obj.name} no existe en el deposito central'
                    },
                    status=400,
                )

            stock_total = (
                stocks_deposito.aggregate(total=Sum('mat_qty'))['total'] or 0
            )

            if stock_total < cantidad:
                return JsonResponse(
                    {
                        'error': f'Stock insuficiente en deposito central. Disponible: {stock_total}'
                    },
                    status=400,
                )

            stock_a_usar = stocks_deposito.order_by('-mat_qty').first()
            stock_a_usar.mat_qty = stock_a_usar.mat_qty - cantidad
            stock_a_usar.save()

            stock_escuela, created = stock_inv.objects.get_or_create(
                inv=escuela,
                mat=material_obj,
                defaults={'mat_qty': cantidad, 'mat_min_qty': 5},
            )

            if not created:
                stock_escuela.mat_qty = stock_escuela.mat_qty + cantidad
                stock_escuela.save()

            movimiento = Movimientos.objects.create(
                tipo='Salida',
                cant_mov=cantidad,
                user=user,
                material=material_obj,
                inv_orig=deposito_central,
                inv_dest=escuela,
                stock_afectado=stock_a_usar,
            )

            return JsonResponse(
                {
                    'success': True,
                    'message': f'Solicitud creada correctamente. Se agregaron {cantidad} {material_obj.get_unidad_display()} a tu inventario.',
                    'movimiento_id': movimiento.id,
                }
            )

    except Exception as e:
        print(f"Error en api_crear_solicitud: {str(e)}")
        import traceback

        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_materiales_escuela(request):
    """Endpoint para obtener materiales disponibles para la escuela"""
    try:
        user = request.user

        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ESCUELA:
            return JsonResponse({'error': 'No autorizado'}, status=403)

        escuela = user.admin_of
        if not escuela:
            return JsonResponse({'data': []})

        materiales_qs = material.objects.all()
        deposito_central = Inventarios.objects.filter(es_depo=True).first()

        data = []
        for m in materiales_qs:
            try:
                stock_deposito = stock_inv.objects.get(inv__es_depo=True, mat=m)
                stock_central = stock_deposito.mat_qty
            except stock_inv.DoesNotExist:
                stock_central = 0

            try:
                stock_escuela = stock_inv.objects.get(inv=escuela, mat=m)
                stock_actual = stock_escuela.mat_qty
                minimo = stock_escuela.mat_min_qty
            except stock_inv.DoesNotExist:
                stock_actual = 0
                minimo = 5

            data.append(
                {
                    'id': m.id,
                    'name': m.name,
                    'tipo': m.get_tipo_display(),
                    'unidad': m.get_unidad_display(),
                    'stock_actual': stock_actual,
                    'stock_central': stock_central,
                    'minimo': minimo,
                    'necesita': max(0, minimo - stock_actual)
                    if stock_actual < minimo
                    else 0,
                }
            )

        return JsonResponse({'data': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ============================================
# REPORTES PARA ADMIN - AÑADIR AL FINAL DE inventario/views.py
# ============================================

@login_required
def api_reporte_inventario_admin(request):
    """Reporte de inventario general para admin"""
    try:
        user = request.user

        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ADMIN:
            return JsonResponse({'error': 'No autorizado'}, status=403)

        escuelas = Inventarios.objects.filter(es_depo=False)

        data = []
        stats_generales = {
            'total_materiales': 0,
            'total_items': 0,
            'bajo_stock': 0,
            'agotados': 0,
            'total_escuelas': escuelas.count(),
        }

        for escuela in escuelas:
            stocks = stock_inv.objects.filter(inv=escuela).select_related('mat')

            for s in stocks:
                estado = 'Normal'
                if s.mat_qty == 0:
                    estado = 'Agotado'
                    stats_generales['agotados'] += 1
                elif s.mat_qty < s.mat_min_qty:
                    estado = 'Bajo Stock'
                    stats_generales['bajo_stock'] += 1

                data.append(
                    {
                        'escuela': escuela.name,
                        'material': s.mat.name,
                        'tipo': s.mat.get_tipo_display(),
                        'unidad': s.mat.get_unidad_display(),
                        'cantidad': s.mat_qty,
                        'minimo': s.mat_min_qty,
                        'estado': estado,
                    }
                )

            stats_generales['total_materiales'] += stocks.count()
            stats_generales['total_items'] += (
                stocks.aggregate(total=Sum('mat_qty'))['total'] or 0
            )

        return JsonResponse(
            {
                'success': True,
                'estadisticas': stats_generales,
                'data': data,
                'fecha': timezone.now().isoformat(),
            }
        )

    except Exception as e:
        print(f"Error en api_reporte_inventario_admin: {str(e)}")
        import traceback

        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_reporte_consumo_admin(request):
    """Reporte de consumo general para admin"""
    try:
        user = request.user

        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ADMIN:
            return JsonResponse({'error': 'No autorizado'}, status=403)

        periodo = request.GET.get('periodo', 'mes')

        hoy = timezone.now()
        if periodo == 'mes':
            fecha_inicio = hoy - timedelta(days=30)
        elif periodo == 'trimestre':
            fecha_inicio = hoy - timedelta(days=90)
        elif periodo == 'semestre':
            fecha_inicio = hoy - timedelta(days=180)
        elif periodo == 'anual':
            fecha_inicio = hoy - timedelta(days=365)
        else:
            fecha_inicio = hoy - timedelta(days=30)

        movimientos = Movimientos.objects.filter(
            tipo='Salida', fecha_Mov__gte=fecha_inicio
        ).select_related('material', 'inv_orig')

        consumo_por_material = {}
        consumo_por_escuela = {}

        for m in movimientos:
            key_mat = m.material_id
            if key_mat not in consumo_por_material:
                consumo_por_material[key_mat] = {
                    'material': m.material.name,
                    'tipo': m.material.get_tipo_display(),
                    'unidad': m.material.get_unidad_display(),
                    'total_consumido': 0,
                    'veces': 0,
                }
            consumo_por_material[key_mat]['total_consumido'] += m.cant_mov
            consumo_por_material[key_mat]['veces'] += 1

            if m.inv_orig:
                key_esc = m.inv_orig_id
                if key_esc not in consumo_por_escuela:
                    consumo_por_escuela[key_esc] = {
                        'escuela': m.inv_orig.name,
                        'total_consumido': 0,
                        'movimientos': 0,
                    }
                consumo_por_escuela[key_esc]['total_consumido'] += m.cant_mov
                consumo_por_escuela[key_esc]['movimientos'] += 1

        stats = {
            'total_movimientos': movimientos.count(),
            'total_consumido': movimientos.aggregate(total=Sum('cant_mov'))['total']
            or 0,
            'periodo': periodo,
            'fecha_inicio': fecha_inicio.isoformat(),
            'fecha_fin': hoy.isoformat(),
        }

        return JsonResponse(
            {
                'success': True,
                'estadisticas': stats,
                'por_material': list(consumo_por_material.values()),
                'por_escuela': list(consumo_por_escuela.values()),
            }
        )

    except Exception as e:
        print(f"Error en api_reporte_consumo_admin: {str(e)}")
        import traceback

        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_reporte_pedidos_admin(request):
    """Reporte de pedidos para admin"""
    try:
        user = request.user

        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ADMIN:
            return JsonResponse({'error': 'No autorizado'}, status=403)

        periodo = request.GET.get('periodo', 'mes')
        estado = request.GET.get('estado', '')

        hoy = timezone.now()
        if periodo == 'mes':
            fecha_inicio = hoy - timedelta(days=30)
        elif periodo == 'trimestre':
            fecha_inicio = hoy - timedelta(days=90)
        elif periodo == 'semestre':
            fecha_inicio = hoy - timedelta(days=180)
        elif periodo == 'anual':
            fecha_inicio = hoy - timedelta(days=365)
        else:
            fecha_inicio = hoy - timedelta(days=30)

        pedidos = Movimientos.objects.filter(
            tipo='Salida', fecha_Mov__gte=fecha_inicio
        ).select_related('material', 'user', 'inv_orig', 'inv_dest')

        if estado:
            pedidos = pedidos.filter(estado=estado)

        data = []
        for p in pedidos:
            estado_pedido = getattr(p, 'estado', 'pendiente')
            data.append(
                {
                    'id': p.id,
                    'fecha': p.fecha_Mov.isoformat(),
                    'material': p.material.name,
                    'cantidad': p.cant_mov,
                    'unidad': p.material.get_unidad_display(),
                    'escuela': p.inv_dest.name if p.inv_dest else 'N/A',
                    'origen': p.inv_orig.name if p.inv_orig else 'Deposito',
                    'solicitante': p.user.username if p.user else 'Sistema',
                    'estado': estado_pedido,
                }
            )

        stats = {
            'total_pedidos': pedidos.count(),
            'total_items': pedidos.aggregate(total=Sum('cant_mov'))['total'] or 0,
            'pendientes': pedidos.filter(estado='pendiente').count(),
            'aprobados': pedidos.filter(estado='aprobado').count(),
            'rechazados': pedidos.filter(estado='rechazado').count(),
            'entregados': pedidos.filter(estado='entregado').count(),
        }

        return JsonResponse({'success': True, 'estadisticas': stats, 'data': data})

    except Exception as e:
        print(f"Error en api_reporte_pedidos_admin: {str(e)}")
        import traceback

        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_reporte_alertas_admin(request):
    """Reporte de alertas para admin"""
    try:
        user = request.user

        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ADMIN:
            return JsonResponse({'error': 'No autorizado'}, status=403)

        alertas = AlertaStock.objects.all().select_related('mat', 'inv').order_by(
            '-fecha'
        )

        data = []
        for a in alertas:
            data.append(
                {
                    'id': a.id,
                    'fecha': a.fecha.isoformat(),
                    'escuela': a.inv.name if a.inv else 'Desconocida',
                    'material': a.mat.name if a.mat else 'Desconocido',
                    'mensaje': a.msg,
                    'resuelto': a.resuelto,
                    'estado': 'Resuelta' if a.resuelto else 'Pendiente',
                }
            )

        stats = {
            'total_alertas': alertas.count(),
            'pendientes': alertas.filter(resuelto=False).count(),
            'resueltas': alertas.filter(resuelto=True).count(),
            'por_escuela': list(
                alertas.values('inv__name').annotate(total=Count('id')).order_by(
                    '-total'
                )
            ),
        }

        return JsonResponse({'success': True, 'estadisticas': stats, 'data': data})

    except Exception as e:
        print(f"Error en api_reporte_alertas_admin: {str(e)}")
        import traceback

        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_exportar_reporte_admin_excel(request):
    """Exportar reporte de admin a Excel"""
    try:
        user = request.user

        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ADMIN:
            return JsonResponse({'error': 'No autorizado'}, status=403)

        tipo = request.GET.get('tipo', 'inventario')
        periodo = request.GET.get('periodo', 'mes')

        wb = Workbook()
        ws = wb.active
        ws.title = f"Reporte {tipo.capitalize()}"

        titulo_font = Font(bold=True, size=14, color="FFFFFF")
        titulo_fill = PatternFill(
            start_color="2C3E50", end_color="2C3E50", fill_type="solid"
        )
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(
            start_color="3498DB", end_color="3498DB", fill_type="solid"
        )
        border = Border(
            left=Side(style='thin', color="DDDDDD"),
            right=Side(style='thin', color="DDDDDD"),
            top=Side(style='thin', color="DDDDDD"),
            bottom=Side(style='thin', color="DDDDDD"),
        )

        ws.merge_cells('A1:E1')
        titulo = ws['A1']
        titulo.value = f"Reporte General de {tipo.capitalize()} - Sistema"
        titulo.font = titulo_font
        titulo.fill = titulo_fill
        titulo.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:E2')
        fecha = ws['A2']
        fecha.value = f"Fecha de generacion: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        fecha.font = Font(italic=True)
        fecha.alignment = Alignment(horizontal='center')

        headers = []
        data_rows = []

        if tipo == 'inventario':
            headers = [
                'Escuela',
                'Material',
                'Tipo',
                'Unidad',
                'Cantidad',
                'Minimo',
                'Estado',
            ]
            escuelas = Inventarios.objects.filter(es_depo=False)
            for escuela in escuelas:
                stocks = stock_inv.objects.filter(inv=escuela).select_related('mat')
                for s in stocks:
                    estado = (
                        'Agotado'
                        if s.mat_qty == 0
                        else (
                            'Bajo Stock'
                            if s.mat_qty < s.mat_min_qty
                            else 'Normal'
                        )
                    )
                    data_rows.append(
                        [
                            escuela.name,
                            s.mat.name,
                            s.mat.get_tipo_display(),
                            s.mat.get_unidad_display(),
                            s.mat_qty,
                            s.mat_min_qty,
                            estado,
                        ]
                    )

        elif tipo == 'consumo':
            headers = ['Escuela', 'Material', 'Total Consumido', 'Veces']
            hoy = timezone.now()
            if periodo == 'mes':
                fecha_inicio = hoy - timedelta(days=30)
            elif periodo == 'trimestre':
                fecha_inicio = hoy - timedelta(days=90)
            elif periodo == 'semestre':
                fecha_inicio = hoy - timedelta(days=180)
            else:
                fecha_inicio = hoy - timedelta(days=365)

            movimientos = Movimientos.objects.filter(
                tipo='Salida', fecha_Mov__gte=fecha_inicio
            ).select_related('material', 'inv_orig')

            consumo_dict = {}
            for m in movimientos:
                key = f"{m.inv_orig_id}_{m.material_id}"
                if key not in consumo_dict:
                    consumo_dict[key] = {
                        'escuela': m.inv_orig.name
                        if m.inv_orig
                        else 'Desconocida',
                        'material': m.material.name,
                        'total': 0,
                        'veces': 0,
                    }
                consumo_dict[key]['total'] += m.cant_mov
                consumo_dict[key]['veces'] += 1

            for item in consumo_dict.values():
                data_rows.append(
                    [item['escuela'], item['material'], item['total'], item['veces']]
                )

        elif tipo == 'pedidos':
            headers = [
                'Fecha',
                'Escuela',
                'Material',
                'Cantidad',
                'Origen',
                'Solicitante',
                'Estado',
            ]
            hoy = timezone.now()
            if periodo == 'mes':
                fecha_inicio = hoy - timedelta(days=30)
            elif periodo == 'trimestre':
                fecha_inicio = hoy - timedelta(days=90)
            elif periodo == 'semestre':
                fecha_inicio = hoy - timedelta(days=180)
            else:
                fecha_inicio = hoy - timedelta(days=365)

            pedidos = Movimientos.objects.filter(
                tipo='Salida', fecha_Mov__gte=fecha_inicio
            ).select_related('material', 'user', 'inv_orig', 'inv_dest')

            for p in pedidos:
                estado_pedido = getattr(p, 'estado', 'pendiente')
                data_rows.append(
                    [
                        p.fecha_Mov.strftime('%d/%m/%Y %H:%M'),
                        p.inv_dest.name if p.inv_dest else 'N/A',
                        p.material.name,
                        p.cant_mov,
                        p.inv_orig.name if p.inv_orig else 'Deposito',
                        p.user.username if p.user else 'Sistema',
                        estado_pedido,
                    ]
                )

        elif tipo == 'alertas':
            headers = ['Fecha', 'Escuela', 'Material', 'Mensaje', 'Estado']
            alertas = AlertaStock.objects.all().select_related('mat', 'inv').order_by(
                '-fecha'
            )

            for a in alertas:
                data_rows.append(
                    [
                        a.fecha.strftime('%d/%m/%Y %H:%M'),
                        a.inv.name if a.inv else 'Desconocida',
                        a.mat.name if a.mat else 'Desconocido',
                        a.msg,
                        'Resuelta' if a.resuelto else 'Pendiente',
                    ]
                )

        row_num = 4
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for row_data in data_rows:
            row_num += 1
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border

        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            for row_idx in range(4, row_num + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    try:
                        max_length = max(max_length, len(str(cell_value)))
                    except:
                        pass
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[chr(64 + col_idx)].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"reporte_admin_{tipo}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    except Exception as e:
        print(f"Error exportando Excel: {str(e)}")
        import traceback

        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)