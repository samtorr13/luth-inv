# transacciones/views.py
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.db import models, transaction
from django.core.exceptions import ValidationError
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import get_user_model
from django.db.models import Sum, Count, Avg, Q, F
from django.utils import timezone
from datetime import timedelta, datetime
import json
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from materiales.models import material
from inventarios.models import Inventarios, AlertaStock, stock_inv
from transacciones.models import Movimientos

User = get_user_model()

# ============================================
# API PARA ESCUELA - MATERIALES DISPONIBLES
# ============================================

@login_required
def api_materiales_escuela(request):
    """Endpoint para obtener materiales disponibles para la escuela"""
    try:
        user = request.user
        print(f"Usuario: {user.username}, Rol: {getattr(user, 'rol', 'No tiene rol')}")
        
        escuela = getattr(user, 'admin_of', None)
        print(f"Escuela: {escuela}")
        
        # Si no tiene escuela, devolver todos los materiales sin stock
        if not escuela:
            materiales_qs = material.objects.all()
            data = []
            for m in materiales_qs:
                data.append({
                    'id': m.id,
                    'name': m.name,
                    'tipo': m.get_tipo_display(),
                    'unidad': m.get_unidad_display(),
                    'stock_actual': 0,
                    'stock_central': 0,
                    'minimo': 5,
                    'necesita': 0
                })
            return JsonResponse({'data': data})
        
        # Obtener todos los materiales
        materiales_qs = material.objects.all()
        
        # Buscar depósito central
        deposito_central = Inventarios.objects.filter(es_depo=True).first()
        
        data = []
        for m in materiales_qs:
            # Stock en depósito central
            stock_central = 0
            if deposito_central:
                stocks_deposito = stock_inv.objects.filter(inv=deposito_central, mat=m)
                if stocks_deposito.exists():
                    stock_central = stocks_deposito.aggregate(total=models.Sum('mat_qty'))['total'] or 0
            
            # Stock en la escuela
            stock_actual = 0
            minimo = 5
            if escuela:
                stocks_escuela = stock_inv.objects.filter(inv=escuela, mat=m)
                if stocks_escuela.exists():
                    stock_actual = stocks_escuela.aggregate(total=models.Sum('mat_qty'))['total'] or 0
                    primer_stock = stocks_escuela.first()
                    minimo = primer_stock.mat_min_qty if primer_stock else 5
            
            necesita = max(0, minimo - stock_actual) if stock_actual < minimo else 0
            
            data.append({
                'id': m.id,
                'name': m.name,
                'tipo': m.get_tipo_display(),
                'unidad': m.get_unidad_display(),
                'stock_actual': stock_actual,
                'stock_central': stock_central,
                'minimo': minimo,
                'necesita': necesita
            })
        
        return JsonResponse({'data': data})
        
    except Exception as e:
        print(f"Error en api_materiales_escuela: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)



# ============================================
# REPORTES PARA ADMIN - BASADOS EN TRANSACCIONES
# ============================================

@login_required
def api_reporte_inventario_admin(request):
    """Reporte de inventario general para admin"""
    try:
        user = request.user
        
        # Verificar que sea admin
        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ADMIN:
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        # Obtener todas las escuelas
        escuelas = Inventarios.objects.filter(es_depo=False)
        
        data = []
        stats_generales = {
            'total_materiales': 0,
            'total_items': 0,
            'bajo_stock': 0,
            'agotados': 0,
            'total_escuelas': escuelas.count()
        }
        
        for escuela in escuelas:
            stocks = stock_inv.objects.filter(inv=escuela).select_related('mat')
            
            for s in stocks:
                estado = 'Normal'
                if s.mat_qty == 0:
                    estado = 'Agotado'
                    stats_generales['agotados'] += 1
                elif s.mat_qty < s.mat_min_qty:
                    estado = 'Bajo Stock'
                    stats_generales['bajo_stock'] += 1
                
                data.append({
                    'escuela': escuela.name,
                    'material': s.mat.name,
                    'tipo': s.mat.get_tipo_display(),
                    'unidad': s.mat.get_unidad_display(),
                    'cantidad': s.mat_qty,
                    'minimo': s.mat_min_qty,
                    'estado': estado
                })
            
            stats_generales['total_materiales'] += stocks.count()
            stats_generales['total_items'] += stocks.aggregate(total=Sum('mat_qty'))['total'] or 0
        
        return JsonResponse({
            'success': True,
            'estadisticas': stats_generales,
            'data': data,
            'fecha': timezone.now().isoformat()
        })
        
    except Exception as e:
        print(f"Error en api_reporte_inventario_admin: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_reporte_consumo_admin(request):
    """Reporte de consumo general para admin"""
    try:
        user = request.user
        
        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ADMIN:
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        periodo = request.GET.get('periodo', 'mes')
        
        # Calcular fecha de inicio
        hoy = timezone.now()
        if periodo == 'mes':
            fecha_inicio = hoy - timedelta(days=30)
        elif periodo == 'trimestre':
            fecha_inicio = hoy - timedelta(days=90)
        elif periodo == 'semestre':
            fecha_inicio = hoy - timedelta(days=180)
        elif periodo == 'anual':
            fecha_inicio = hoy - timedelta(days=365)
        else:
            fecha_inicio = hoy - timedelta(days=30)
        
        # Obtener todos los movimientos de salida
        movimientos = Movimientos.objects.filter(
            tipo='Salida',
            fecha_Mov__gte=fecha_inicio
        ).select_related('material', 'inv_orig')
        
        # Agrupar por material
        consumo_por_material = {}
        # Agrupar por escuela
        consumo_por_escuela = {}
        
        for m in movimientos:
            # Por material
            key_mat = m.material_id
            if key_mat not in consumo_por_material:
                consumo_por_material[key_mat] = {
                    'material': m.material.name,
                    'tipo': m.material.get_tipo_display(),
                    'unidad': m.material.get_unidad_display(),
                    'total_consumido': 0,
                    'veces': 0
                }
            consumo_por_material[key_mat]['total_consumido'] += m.cant_mov
            consumo_por_material[key_mat]['veces'] += 1
            
            # Por escuela
            if m.inv_orig:
                key_esc = m.inv_orig_id
                if key_esc not in consumo_por_escuela:
                    consumo_por_escuela[key_esc] = {
                        'escuela': m.inv_orig.name,
                        'total_consumido': 0,
                        'movimientos': 0
                    }
                consumo_por_escuela[key_esc]['total_consumido'] += m.cant_mov
                consumo_por_escuela[key_esc]['movimientos'] += 1
        
        stats = {
            'total_movimientos': movimientos.count(),
            'total_consumido': movimientos.aggregate(total=Sum('cant_mov'))['total'] or 0,
            'periodo': periodo,
            'fecha_inicio': fecha_inicio.isoformat(),
            'fecha_fin': hoy.isoformat()
        }
        
        return JsonResponse({
            'success': True,
            'estadisticas': stats,
            'por_material': list(consumo_por_material.values()),
            'por_escuela': list(consumo_por_escuela.values())
        })
        
    except Exception as e:
        print(f"Error en api_reporte_consumo_admin: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_reporte_pedidos_admin(request):
    """Reporte de pedidos para admin"""
    try:
        user = request.user
        
        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ADMIN:
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        periodo = request.GET.get('periodo', 'mes')
        estado = request.GET.get('estado', '')
        
        hoy = timezone.now()
        if periodo == 'mes':
            fecha_inicio = hoy - timedelta(days=30)
        elif periodo == 'trimestre':
            fecha_inicio = hoy - timedelta(days=90)
        elif periodo == 'semestre':
            fecha_inicio = hoy - timedelta(days=180)
        elif periodo == 'anual':
            fecha_inicio = hoy - timedelta(days=365)
        else:
            fecha_inicio = hoy - timedelta(days=30)
        
        # Obtener movimientos de salida (pedidos)
        pedidos = Movimientos.objects.filter(
            tipo='Salida',
            fecha_Mov__gte=fecha_inicio
        ).select_related('material', 'user', 'inv_orig', 'inv_dest')
        
        if estado:
            pedidos = pedidos.filter(estado=estado)
        
        data = []
        for p in pedidos:
            estado_pedido = getattr(p, 'estado', 'pendiente')
            data.append({
                'id': p.id,
                'fecha': p.fecha_Mov.isoformat(),
                'material': p.material.name,
                'cantidad': p.cant_mov,
                'unidad': p.material.get_unidad_display(),
                'escuela': p.inv_dest.name if p.inv_dest else 'N/A',
                'origen': p.inv_orig.name if p.inv_orig else 'Depósito',
                'solicitante': p.user.username if p.user else 'Sistema',
                'estado': estado_pedido
            })
        
        stats = {
            'total_pedidos': pedidos.count(),
            'total_items': pedidos.aggregate(total=Sum('cant_mov'))['total'] or 0,
            'pendientes': pedidos.filter(estado='pendiente').count(),
            'aprobados': pedidos.filter(estado='aprobado').count(),
            'rechazados': pedidos.filter(estado='rechazado').count(),
            'entregados': pedidos.filter(estado='entregado').count()
        }
        
        return JsonResponse({
            'success': True,
            'estadisticas': stats,
            'data': data
        })
        
    except Exception as e:
        print(f"Error en api_reporte_pedidos_admin: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_reporte_alertas_admin(request):
    """Reporte de alertas para admin"""
    try:
        user = request.user
        
        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ADMIN:
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        # Obtener todas las alertas
        alertas = AlertaStock.objects.all().select_related('mat', 'inv').order_by('-fecha')
        
        data = []
        for a in alertas:
            data.append({
                'id': a.id,
                'fecha': a.fecha.isoformat(),
                'escuela': a.inv.name if a.inv else 'Desconocida',
                'material': a.mat.name if a.mat else 'Desconocido',
                'mensaje': a.msg,
                'resuelto': a.resuelto,
                'estado': 'Resuelta' if a.resuelto else 'Pendiente'
            })
        
        stats = {
            'total_alertas': alertas.count(),
            'pendientes': alertas.filter(resuelto=False).count(),
            'resueltas': alertas.filter(resuelto=True).count(),
            'por_escuela': alertas.values('inv__name').annotate(total=Count('id')).order_by('-total')
        }
        
        return JsonResponse({
            'success': True,
            'estadisticas': stats,
            'data': data
        })
        
    except Exception as e:
        print(f"Error en api_reporte_alertas_admin: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_exportar_reporte_admin_excel(request):
    """Exportar reporte de admin a Excel"""
    try:
        user = request.user
        
        if not hasattr(user, 'rol') or user.rol != CustomUser.Rol.ADMIN:
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        tipo = request.GET.get('tipo', 'inventario')
        periodo = request.GET.get('periodo', 'mes')
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Reporte {tipo.capitalize()}"
        
        # Estilos
        titulo_font = Font(bold=True, size=14, color="FFFFFF")
        titulo_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        border = Border(
            left=Side(style='thin', color="DDDDDD"),
            right=Side(style='thin', color="DDDDDD"),
            top=Side(style='thin', color="DDDDDD"),
            bottom=Side(style='thin', color="DDDDDD")
        )
        
        # Título
        ws.merge_cells('A1:E1')
        titulo = ws['A1']
        titulo.value = f"Reporte General de {tipo.capitalize()} - Sistema"
        titulo.font = titulo_font
        titulo.fill = titulo_fill
        titulo.alignment = Alignment(horizontal='center', vertical='center')
        
        # Fecha
        ws.merge_cells('A2:E2')
        fecha = ws['A2']
        fecha.value = f"Fecha de generación: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        fecha.font = Font(italic=True)
        fecha.alignment = Alignment(horizontal='center')
        
        # Obtener datos según tipo
        headers = []
        data_rows = []
        
        if tipo == 'inventario':
            headers = ['Escuela', 'Material', 'Tipo', 'Unidad', 'Cantidad', 'Mínimo', 'Estado']
            escuelas = Inventarios.objects.filter(es_depo=False)
            for escuela in escuelas:
                stocks = stock_inv.objects.filter(inv=escuela).select_related('mat')
                for s in stocks:
                    estado = 'Agotado' if s.mat_qty == 0 else ('Bajo Stock' if s.mat_qty < s.mat_min_qty else 'Normal')
                    data_rows.append([
                        escuela.name,
                        s.mat.name,
                        s.mat.get_tipo_display(),
                        s.mat.get_unidad_display(),
                        s.mat_qty,
                        s.mat_min_qty,
                        estado
                    ])
        
        elif tipo == 'consumo':
            headers = ['Escuela', 'Material', 'Total Consumido', 'Veces']
            hoy = timezone.now()
            if periodo == 'mes':
                fecha_inicio = hoy - timedelta(days=30)
            elif periodo == 'trimestre':
                fecha_inicio = hoy - timedelta(days=90)
            elif periodo == 'semestre':
                fecha_inicio = hoy - timedelta(days=180)
            else:
                fecha_inicio = hoy - timedelta(days=365)
            
            movimientos = Movimientos.objects.filter(
                tipo='Salida',
                fecha_Mov__gte=fecha_inicio
            ).select_related('material', 'inv_orig')
            
            consumo_dict = {}
            for m in movimientos:
                key = f"{m.inv_orig_id}_{m.material_id}"
                if key not in consumo_dict:
                    consumo_dict[key] = {
                        'escuela': m.inv_orig.name if m.inv_orig else 'Desconocida',
                        'material': m.material.name,
                        'total': 0,
                        'veces': 0
                    }
                consumo_dict[key]['total'] += m.cant_mov
                consumo_dict[key]['veces'] += 1
            
            for item in consumo_dict.values():
                data_rows.append([
                    item['escuela'],
                    item['material'],
                    item['total'],
                    item['veces']
                ])
        
        elif tipo == 'pedidos':
            headers = ['Fecha', 'Escuela', 'Material', 'Cantidad', 'Origen', 'Solicitante', 'Estado']
            hoy = timezone.now()
            if periodo == 'mes':
                fecha_inicio = hoy - timedelta(days=30)
            elif periodo == 'trimestre':
                fecha_inicio = hoy - timedelta(days=90)
            elif periodo == 'semestre':
                fecha_inicio = hoy - timedelta(days=180)
            else:
                fecha_inicio = hoy - timedelta(days=365)
            
            pedidos = Movimientos.objects.filter(
                tipo='Salida',
                fecha_Mov__gte=fecha_inicio
            ).select_related('material', 'user', 'inv_orig', 'inv_dest')
            
            for p in pedidos:
                estado_pedido = getattr(p, 'estado', 'pendiente')
                data_rows.append([
                    p.fecha_Mov.strftime('%d/%m/%Y %H:%M'),
                    p.inv_dest.name if p.inv_dest else 'N/A',
                    p.material.name,
                    p.cant_mov,
                    p.inv_orig.name if p.inv_orig else 'Depósito',
                    p.user.username if p.user else 'Sistema',
                    estado_pedido
                ])
        
        elif tipo == 'alertas':
            headers = ['Fecha', 'Escuela', 'Material', 'Mensaje', 'Estado']
            alertas = AlertaStock.objects.all().select_related('mat', 'inv').order_by('-fecha')
            
            for a in alertas:
                data_rows.append([
                    a.fecha.strftime('%d/%m/%Y %H:%M'),
                    a.inv.name if a.inv else 'Desconocida',
                    a.mat.name if a.mat else 'Desconocido',
                    a.msg,
                    'Resuelta' if a.resuelto else 'Pendiente'
                ])
        
        # Escribir headers
        row_num = 4
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Escribir datos
        for row_data in data_rows:
            row_num += 1
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
        
        # Ajustar ancho de columnas
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            for row_idx in range(4, row_num + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    try:
                        max_length = max(max_length, len(str(cell_value)))
                    except:
                        pass
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[chr(64 + col_idx)].width = adjusted_width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"reporte_admin_{tipo}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        print(f"Error exportando Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)
# ============================================
# API PARA ESCUELA - MOVIMIENTOS/PEDIDOS
# ============================================

@login_required
def api_movimientos_escuela(request):
    """Endpoint para obtener movimientos/pedidos de la escuela del usuario"""
    try:
        user = request.user
        
        escuela = getattr(user, 'admin_of', None)
        if not escuela:
            return JsonResponse({'data': []})
        
        # Obtener movimientos donde esta escuela sea origen o destino
        movimientos = Movimientos.objects.filter(
            models.Q(inv_orig=escuela) | models.Q(inv_dest=escuela)
        ).select_related('material', 'user').order_by('-fecha_Mov')[:50]
        
        data = []
        for m in movimientos:
            data.append({
                'id': m.id,
                'material_id': m.material.id,
                'material_nombre': m.material.name,
                'cantidad': m.cant_mov,
                'tipo': m.tipo,
                'fecha': m.fecha_Mov.isoformat(),
                'origen': m.inv_orig.name if m.inv_orig else None,
                'destino': m.inv_dest.name if m.inv_dest else None,
                'usuario': m.user.username if m.user else 'Sistema',
                'es_salida': m.tipo == 'Salida',
                'es_entrada': m.tipo == 'Entrada',
                'es_ajuste': m.tipo == 'Ajuste'
            })
        
        return JsonResponse({'data': data})
        
    except Exception as e:
        print(f"Error en api_movimientos_escuela: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


# ============================================
# API PARA ESCUELA - CREAR SOLICITUD
# ============================================

@login_required
@csrf_exempt
def api_crear_solicitud(request):
    """Endpoint para crear una solicitud de material (OUT + IN)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        user = request.user
        data = json.loads(request.body)
        
        material_id = data.get('material_id')
        cantidad = data.get('cantidad')
        notas = data.get('notas', '')
        
        # Validar datos
        if not material_id or not cantidad:
            return JsonResponse({'error': 'Material y cantidad son requeridos'}, status=400)
        
        try:
            material_obj = material.objects.get(id=material_id)
        except material.DoesNotExist:
            return JsonResponse({'error': 'Material no encontrado'}, status=404)
        
        escuela = getattr(user, 'admin_of', None)
        if not escuela:
            return JsonResponse({'error': 'No tienes una escuela asignada'}, status=400)
        
        # Buscar depósito central
        deposito_central = Inventarios.objects.filter(es_depo=True).first()
        if not deposito_central:
            return JsonResponse({'error': 'No hay depósito central configurado'}, status=400)
        
        with transaction.atomic():
            # Verificar stock en depósito central
            stocks_deposito = stock_inv.objects.filter(inv=deposito_central, mat=material_obj)
            
            if not stocks_deposito.exists():
                return JsonResponse({
                    'error': f'El material {material_obj.name} no existe en el depósito central'
                }, status=400)
            
            # Calcular stock total
            stock_total = stocks_deposito.aggregate(total=models.Sum('mat_qty'))['total'] or 0
            
            if stock_total < cantidad:
                return JsonResponse({
                    'error': f'Stock insuficiente en depósito central. Disponible: {stock_total}'
                }, status=400)
            
            # Seleccionar el stock con mayor cantidad para restar
            stock_a_usar = stocks_deposito.order_by('-mat_qty').first()
            
            # RESTAR del depósito central
            stock_a_usar.mat_qty = stock_a_usar.mat_qty - cantidad
            stock_a_usar.save()
            
            # SUMAR a la escuela (crear o actualizar)
            stock_escuela, created = stock_inv.objects.get_or_create(
                inv=escuela,
                mat=material_obj,
                defaults={'mat_qty': cantidad, 'mat_min_qty': 5}
            )
            
            if not created:
                stock_escuela.mat_qty = stock_escuela.mat_qty + cantidad
                stock_escuela.save()
            
            # Registrar el movimiento para historial
            movimiento = Movimientos.objects.create(
                tipo='Salida',
                cant_mov=cantidad,
                user=user,
                material=material_obj,
                inv_orig=deposito_central,
                inv_dest=escuela,
                stock_afectado=stock_a_usar
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Solicitud creada correctamente. Se agregaron {cantidad} {material_obj.get_unidad_display()} a tu inventario.',
                'movimiento_id': movimiento.id
            })
            
    except Exception as e:
        print(f"Error en api_crear_solicitud: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


# ============================================
# REPORTES PARA ESCUELA
# ============================================

@login_required
def api_reporte_inventario(request):
    """Reporte de inventario actual de la escuela"""
    try:
        user = request.user
        escuela = getattr(user, 'admin_of', None)
        
        if not escuela:
            return JsonResponse({'success': False, 'error': 'No tienes una escuela asignada'}, status=400)
        
        # Obtener stocks de la escuela
        stocks = stock_inv.objects.filter(inv=escuela).select_related('mat')
        
        # Datos del reporte
        data = []
        bajo_stock = 0
        agotados = 0
        
        for s in stocks:
            estado = 'Normal'
            if s.mat_qty == 0:
                estado = 'Agotado'
                agotados += 1
            elif s.mat_qty < s.mat_min_qty:
                estado = 'Bajo Stock'
                bajo_stock += 1
            
            data.append({
                'material': s.mat.name,
                'tipo': s.mat.get_tipo_display(),
                'unidad': s.mat.get_unidad_display(),
                'cantidad': s.mat_qty,
                'minimo': s.mat_min_qty,
                'estado': estado
            })
        
        # Estadísticas
        stats = {
            'total_materiales': stocks.count(),
            'total_items': stocks.aggregate(total=Sum('mat_qty'))['total'] or 0,
            'bajo_stock': bajo_stock,
            'agotados': agotados,
            'fecha_generacion': timezone.now().isoformat()
        }
        
        return JsonResponse({
            'success': True,
            'reporte': 'inventario',
            'estadisticas': stats,
            'data': data
        })
        
    except Exception as e:
        print(f"Error en api_reporte_inventario: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def api_reporte_consumo(request):
    """Reporte de consumo mensual de la escuela"""
    try:
        user = request.user
        escuela = getattr(user, 'admin_of', None)
        periodo = request.GET.get('periodo', 'mes')
        
        if not escuela:
            return JsonResponse({'success': False, 'error': 'No tienes una escuela asignada'}, status=400)
        
        # Calcular fecha de inicio según período
        hoy = timezone.now()
        if periodo == 'mes':
            fecha_inicio = hoy - timedelta(days=30)
        elif periodo == 'trimestre':
            fecha_inicio = hoy - timedelta(days=90)
        elif periodo == 'semestre':
            fecha_inicio = hoy - timedelta(days=180)
        elif periodo == 'anual':
            fecha_inicio = hoy - timedelta(days=365)
        else:
            fecha_inicio = hoy - timedelta(days=30)
        
        # Obtener movimientos de SALIDA de la escuela
        movimientos = Movimientos.objects.filter(
            inv_orig=escuela,
            tipo='Salida',
            fecha_Mov__gte=fecha_inicio
        ).select_related('material')
        
        # Agrupar por material
        consumo_por_material = {}
        for m in movimientos:
            key = m.material_id
            if key not in consumo_por_material:
                consumo_por_material[key] = {
                    'material': m.material.name,
                    'tipo': m.material.get_tipo_display(),
                    'unidad': m.material.get_unidad_display(),
                    'total_consumido': 0,
                    'veces_solicitado': 0
                }
            consumo_por_material[key]['total_consumido'] += m.cant_mov
            consumo_por_material[key]['veces_solicitado'] += 1
        
        # Estadísticas
        stats = {
            'total_movimientos': movimientos.count(),
            'total_consumido': movimientos.aggregate(total=Sum('cant_mov'))['total'] or 0,
            'periodo': periodo,
            'fecha_inicio': fecha_inicio.isoformat(),
            'fecha_fin': hoy.isoformat()
        }
        
        return JsonResponse({
            'success': True,
            'reporte': 'consumo',
            'estadisticas': stats,
            'data': list(consumo_por_material.values())
        })
        
    except Exception as e:
        print(f"Error en api_reporte_consumo: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def api_reporte_pedidos(request):
    """Reporte de historial de pedidos de la escuela"""
    try:
        user = request.user
        escuela = getattr(user, 'admin_of', None)
        periodo = request.GET.get('periodo', 'mes')
        
        if not escuela:
            return JsonResponse({'success': False, 'error': 'No tienes una escuela asignada'}, status=400)
        
        # Calcular fecha de inicio
        hoy = timezone.now()
        if periodo == 'mes':
            fecha_inicio = hoy - timedelta(days=30)
        elif periodo == 'trimestre':
            fecha_inicio = hoy - timedelta(days=90)
        elif periodo == 'semestre':
            fecha_inicio = hoy - timedelta(days=180)
        elif periodo == 'anual':
            fecha_inicio = hoy - timedelta(days=365)
        else:
            fecha_inicio = hoy - timedelta(days=30)
        
        # Obtener movimientos donde la escuela es destino (pedidos recibidos)
        pedidos = Movimientos.objects.filter(
            inv_dest=escuela,
            fecha_Mov__gte=fecha_inicio
        ).select_related('material', 'user', 'inv_orig').order_by('-fecha_Mov')
        
        data = []
        for p in pedidos:
            data.append({
                'id': p.id,
                'fecha': p.fecha_Mov.isoformat(),
                'material': p.material.name,
                'cantidad': p.cant_mov,
                'origen': p.inv_orig.name if p.inv_orig else 'Depósito Central',
                'usuario': p.user.username if p.user else 'Sistema'
            })
        
        # Estadísticas
        stats = {
            'total_pedidos': pedidos.count(),
            'total_items': pedidos.aggregate(total=Sum('cant_mov'))['total'] or 0,
            'periodo': periodo,
            'fecha_inicio': fecha_inicio.isoformat(),
            'fecha_fin': hoy.isoformat()
        }
        
        return JsonResponse({
            'success': True,
            'reporte': 'pedidos',
            'estadisticas': stats,
            'data': data
        })
        
    except Exception as e:
        print(f"Error en api_reporte_pedidos: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def api_reporte_alertas(request):
    """Reporte de alertas de la escuela"""
    try:
        user = request.user
        escuela = getattr(user, 'admin_of', None)
        
        if not escuela:
            return JsonResponse({'success': False, 'error': 'No tienes una escuela asignada'}, status=400)
        
        # Obtener alertas
        alertas = AlertaStock.objects.filter(
            inv=escuela
        ).select_related('mat').order_by('-fecha')[:100]
        
        data = []
        for a in alertas:
            data.append({
                'id': a.id,
                'fecha': a.fecha.isoformat(),
                'material': a.mat.name,
                'mensaje': a.msg,
                'resuelto': a.resuelto,
                'estado': 'Resuelta' if a.resuelto else 'Pendiente'
            })
        
        # Estadísticas
        stats = {
            'total_alertas': alertas.count(),
            'pendientes': alertas.filter(resuelto=False).count(),
            'resueltas': alertas.filter(resuelto=True).count()
        }
        
        return JsonResponse({
            'success': True,
            'reporte': 'alertas',
            'estadisticas': stats,
            'data': data
        })
        
    except Exception as e:
        print(f"Error en api_reporte_alertas: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================
# EXPORTAR REPORTE A EXCEL
# ============================================

@login_required
def api_exportar_reporte_excel(request):
    """Exportar reporte a Excel"""
    try:
        user = request.user
        escuela = getattr(user, 'admin_of', None)
        
        if not escuela:
            return JsonResponse({'error': 'No tienes una escuela asignada'}, status=400)
        
        tipo = request.GET.get('tipo', 'inventario')
        periodo = request.GET.get('periodo', 'mes')
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Reporte {tipo.capitalize()}"
        
        # Estilos
        titulo_font = Font(bold=True, size=14, color="FFFFFF")
        titulo_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        border = Border(
            left=Side(style='thin', color="DDDDDD"),
            right=Side(style='thin', color="DDDDDD"),
            top=Side(style='thin', color="DDDDDD"),
            bottom=Side(style='thin', color="DDDDDD")
        )
        
        # Título
        ws.merge_cells('A1:E1')
        titulo = ws['A1']
        titulo.value = f"Reporte de {tipo.capitalize()} - {escuela.name}"
        titulo.font = titulo_font
        titulo.fill = titulo_fill
        titulo.alignment = Alignment(horizontal='center', vertical='center')
        
        # Fecha
        ws.merge_cells('A2:E2')
        fecha = ws['A2']
        fecha.value = f"Fecha de generación: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        fecha.font = Font(italic=True)
        fecha.alignment = Alignment(horizontal='center')
        
        # Obtener datos según tipo
        headers = []
        data_rows = []
        
        if tipo == 'inventario':
            headers = ['Material', 'Tipo', 'Unidad', 'Cantidad', 'Mínimo', 'Estado']
            stocks = stock_inv.objects.filter(inv=escuela).select_related('mat')
            for s in stocks:
                estado = 'Agotado' if s.mat_qty == 0 else ('Bajo Stock' if s.mat_qty < s.mat_min_qty else 'Normal')
                data_rows.append([
                    s.mat.name,
                    s.mat.get_tipo_display(),
                    s.mat.get_unidad_display(),
                    s.mat_qty,
                    s.mat_min_qty,
                    estado
                ])
        
        elif tipo == 'consumo':
            headers = ['Material', 'Tipo', 'Unidad', 'Total Consumido', 'Veces Solicitado']
            hoy = timezone.now()
            if periodo == 'mes':
                fecha_inicio = hoy - timedelta(days=30)
            elif periodo == 'trimestre':
                fecha_inicio = hoy - timedelta(days=90)
            elif periodo == 'semestre':
                fecha_inicio = hoy - timedelta(days=180)
            else:  # anual
                fecha_inicio = hoy - timedelta(days=365)
            
            movimientos = Movimientos.objects.filter(
                inv_orig=escuela,
                tipo='Salida',
                fecha_Mov__gte=fecha_inicio
            ).select_related('material')
            
            consumo_dict = {}
            for m in movimientos:
                key = m.material_id
                if key not in consumo_dict:
                    consumo_dict[key] = {
                        'nombre': m.material.name,
                        'tipo': m.material.get_tipo_display(),
                        'unidad': m.material.get_unidad_display(),
                        'total': 0,
                        'veces': 0
                    }
                consumo_dict[key]['total'] += m.cant_mov
                consumo_dict[key]['veces'] += 1
            
            for item in consumo_dict.values():
                data_rows.append([
                    item['nombre'],
                    item['tipo'],
                    item['unidad'],
                    item['total'],
                    item['veces']
                ])
        
        elif tipo == 'pedidos':
            headers = ['Fecha', 'Material', 'Cantidad', 'Origen', 'Usuario']
            hoy = timezone.now()
            if periodo == 'mes':
                fecha_inicio = hoy - timedelta(days=30)
            elif periodo == 'trimestre':
                fecha_inicio = hoy - timedelta(days=90)
            elif periodo == 'semestre':
                fecha_inicio = hoy - timedelta(days=180)
            else:
                fecha_inicio = hoy - timedelta(days=365)
            
            pedidos = Movimientos.objects.filter(
                inv_dest=escuela,
                tipo='Salida',
                fecha_Mov__gte=fecha_inicio
            ).select_related('material', 'user').order_by('-fecha_Mov')
            
            for p in pedidos:
                data_rows.append([
                    p.fecha_Mov.strftime('%d/%m/%Y %H:%M'),
                    p.material.name,
                    p.cant_mov,
                    p.inv_orig.name if p.inv_orig else 'Depósito Central',
                    p.user.username if p.user else 'Sistema'
                ])
        
        elif tipo == 'alertas':
            headers = ['Fecha', 'Material', 'Mensaje', 'Estado']
            alertas = AlertaStock.objects.filter(
                inv=escuela
            ).select_related('mat').order_by('-fecha')
            
            for a in alertas:
                data_rows.append([
                    a.fecha.strftime('%d/%m/%Y %H:%M'),
                    a.mat.name,
                    a.msg,
                    'Resuelta' if a.resuelto else 'Pendiente'
                ])
        
        # Escribir headers
        row_num = 4
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Escribir datos
        for row_data in data_rows:
            row_num += 1
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                if col_num == len(headers):  # Última columna (estado)
                    if 'Agotado' in str(value):
                        cell.font = Font(color="E74C3C", bold=True)
                    elif 'Bajo' in str(value):
                        cell.font = Font(color="F39C12", bold=True)
        
        # Ajustar ancho de columnas
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            for row_idx in range(4, row_num + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    try:
                        max_length = max(max_length, len(str(cell_value)))
                    except:
                        pass
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[chr(64 + col_idx)].width = adjusted_width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"reporte_{tipo}_{escuela.name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        print(f"Error exportando Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)